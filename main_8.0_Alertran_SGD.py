import sys
import asyncio
from pathlib import Path
from typing import List, Union
from openpyxl import load_workbook, Workbook
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeoutError, Error as PlaywrightError
from datetime import datetime, timedelta
import time
import os
import subprocess

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QComboBox, QPushButton, QTextEdit, QProgressBar,
    QFileDialog, QMessageBox, QGroupBox, QFormLayout, QFrame, QDialog,
    QDialogButtonBox, QSpinBox, QTableWidget, QTableWidgetItem, QHeaderView,
    QStyleOptionProgressBar, QStyle
)
from PySide6.QtCore import Signal, QObject, Qt, QThread, QEventLoop, QTimer, QPropertyAnimation, QEasingCurve, Property
from PySide6.QtGui import QFont, QIcon, QTextCursor, QPalette, QColor, QPainter, QLinearGradient, QBrush, QPen
import qasync

# =======================
# FUNCIÓN PARA OBTENER CARPETA DE DESCARGAS
# =======================
def obtener_carpeta_descargas():
    """Obtiene la ruta de la carpeta de Descargas del usuario"""
    home = Path.home()
    
    # Sistema operativo Windows
    if os.name == 'nt':
        downloads = home / 'Downloads'
        if downloads.exists():
            return downloads
    
    # Sistema operativo macOS/Linux
    downloads = home / 'Descargas'  # Español
    if downloads.exists():
        return downloads
    
    downloads = home / 'Downloads'  # Inglés
    if downloads.exists():
        return downloads
    
    # Si no encuentra ninguna, crear carpeta en el directorio del usuario
    descargas_dir = home / 'Descargas_Alertran'
    descargas_dir.mkdir(exist_ok=True)
    return descargas_dir

# =======================
# FUNCIÓN PARA GENERAR NOMBRE DE ARCHIVO ÚNICO
# =======================
def generar_nombre_archivo(base_nombre, extension, carpeta):
    """Genera un nombre de archivo único añadiendo un número si ya existe"""
    contador = 1
    nombre_archivo = carpeta / f"{base_nombre}.{extension}"
    
    while nombre_archivo.exists():
        nombre_archivo = carpeta / f"{base_nombre}_{contador}.{extension}"
        contador += 1
    
    return nombre_archivo

# =======================
# CONSTANTES
# =======================
CIUDADES = [
  "ABA BARRANQUILLA AEROPUER","ABG BUCARAMANGA AEROPUER","ADZ SAN ANDRES",
  "AEJ BARRANCABERMEJA AERO","AIB IBAGUE AEROPUERTO","ALM ALAMOS",
  "AMT MONTERIA AEROPUERTO","ANV NEIVA AEROPUERTO","APA PALMIRA AEROPUERTO",
  "APE PEREIRA AEROPUERTO","APO APARTADO","APS PASTO AEROPUERTO",
  "ARC RIOHACHA AEROPUERTO","ARN RIONEGRO AEROPUERTO",
  "ASM SANTA MARTA AEROPUER","ATC TUMACO AEROPUERTO",
  "AVU VALLEDUPAR AEROPUERT","AXM ARMENIA","BAQ BARRANQUILLA",
  "BGA BUCARAMANGA","BGG BUGA","BOG BOGOTA","BUN BUENAVENTURA",
  "CAQ CAUCASIA","CGW CARTAGO","CLO CALI","CNG CIENAGA","CTG CARTAGENA",
  "CUC CUCUTA","CZL COROZAL","DTA DUITAMA","EJA BARRANCABERMEJ",
  "EYP EL YOPAL","FLA FLORENCIA","GIR GIRARDOT","HDA HONDA","IBE IBAGUE",
  "IPI IPIALES","LDR LA DORADA","LET LETICIA","MAQ MARIQUITA","MCO MAICAO",
  "MDE MEDELLIN","MTR MONTERIA","MZL MANIZALES","NVA NEIVA","PAL PAMPLONA",
  "PAM PALMIRA","PEI PEREIRA","PPN POPAYAN","PSO PASTO","PTO PUERTO BOYACA",
  "RCH RIOHACHA","RNG RIONEGRO","RZG REZAGOS","SIJ SINCELEJO","SMR SANTA MARTA",
  "SOX SOGAMOSO","STQ SANTANDER DE QUILICH","TCO TUMACO –","TNB TUNJA","UIB QUIDBO",
  "ULQ TULUA","VUP VALLEDUPAR","WC VILLAVICENCIO","000 CENTRAL",
  "900 LOGISTICA INTERNACIO","901 BODEGA MIAMI","999 TODAS LAS PLAZAS"  
]

TIPOS_INCIDENCIA = ["22","27","50","51","52","920","929"]
MAX_REINTENTOS = 3
TIEMPO_ESPERA_RECUPERACION = 4000
TIEMPO_ESPERA_NAVEGACION = 3000
TIEMPO_ESPERA_CLICK = 2000
TIEMPO_ESPERA_CARGA = 8000
TIEMPO_ESPERA_ENTRE_GUIAS = 2000
TIEMPO_ESPERA_INGRESO_CODIGOS = 1500
TIEMPO_ESPERA_VOLVER = 5000

# =======================
# BARRA DE PROGRESO 
# =======================
class MacProgressBar(QProgressBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(25)
        self.setMaximumHeight(25)
        self.setTextVisible(False)
        self._animation = QPropertyAnimation(self, b"value")
        self._animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        self._animation.setDuration(300)  # Duración de la animación en ms
        
    def setValue(self, value):
        # Animar el cambio de valor
        if self._animation.state() == QPropertyAnimation.State.Running:
            self._animation.stop()
        
        self._animation.setStartValue(self.value())
        self._animation.setEndValue(value)
        self._animation.start()
        super().setValue(value)
    
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)
        
        # Configurar el rectángulo de la barra
        rect = self.rect()
        rect_width = rect.width()
        rect_height = rect.height()
        
        # Dibujar fondo (gris claro)
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(QColor(220, 220, 220)))  # Gris claro
        painter.drawRoundedRect(rect, 8, 8)
        
        # Calcular el ancho del progreso
        progress = self.value() / 100.0
        progress_width = int(rect_width * progress)
        
        if progress_width > 0:
            # Crear gradiente azul estilo Mac
            gradient = QLinearGradient(0, 0, progress_width, 0)
            gradient.setColorAt(0, QColor(10, 132, 255))    # Azul claro
            gradient.setColorAt(0.7, QColor(0, 100, 255))   # Azul medio
            gradient.setColorAt(1, QColor(0, 85, 255))      # Azul oscuro
            
            # Dibujar el progreso con gradiente
            progress_rect = rect.adjusted(0, 0, -(rect_width - progress_width), 0)
            painter.setBrush(QBrush(gradient))
            painter.drawRoundedRect(progress_rect, 8, 8)
            
            # Dibujar brillo superior (efecto glass)
            highlight_rect = progress_rect.adjusted(0, 0, 0, -rect_height//2)
            highlight_gradient = QLinearGradient(0, 0, 0, highlight_rect.height())
            highlight_gradient.setColorAt(0, QColor(255, 255, 255, 70))
            highlight_gradient.setColorAt(1, QColor(255, 255, 255, 20))
            painter.setBrush(QBrush(highlight_gradient))
            painter.drawRoundedRect(highlight_rect, 8, 8)
            
            # Dibujar sombra interior
            shadow_rect = progress_rect.adjusted(0, rect_height//2, 0, 0)
            shadow_gradient = QLinearGradient(0, 0, 0, shadow_rect.height())
            shadow_gradient.setColorAt(0, QColor(0, 0, 0, 20))
            shadow_gradient.setColorAt(1, QColor(0, 0, 0, 5))
            painter.setBrush(QBrush(shadow_gradient))
            painter.drawRoundedRect(shadow_rect, 8, 8)
        
        # Dibujar borde
        painter.setPen(QPen(QColor(150, 150, 150), 1))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRoundedRect(rect, 8, 8)
        
        # Dibujar texto del porcentaje
        if self.value() > 0:
            painter.setPen(QColor(255, 255, 255))
            painter.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            text = f"{self.value()}%"
            text_rect = rect
            painter.drawText(text_rect, Qt.AlignmentFlag.AlignCenter, text)

# =======================
# VENTANA DE RESUMEN
# =======================
class ResumenWindow(QDialog):
    def __init__(self, total_guias, desviadas, entregadas, errores, advertencias, tiempo_total, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📊 RESUMEN DEL PROCESO")
        self.setMinimumWidth(500)
        self.setMinimumHeight(400)
        self.setModal(True)
        
        # Estilo de la ventana
        self.setStyleSheet("""
            QDialog {
                rgba(255, 255, 255, 0.95);
            }
            QLabel {
                font-size: 12pt;
                padding: 5px;
            }
            .titulo {
                font-size: 18pt;
                font-weight: bold;
                color: #2c3e50;
            }
            .numero {
                font-size: 24pt;
                font-weight: bold;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Título
        titulo = QLabel("✅ PROCESO COMPLETADO")
        titulo.setProperty("class", "titulo")
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(titulo)
        
        # Icono de éxito
        icono = QLabel("")
        icono.setFont(QFont("Arial", 48))
        icono.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(icono)
        
        # Línea separadora
        linea = QFrame()
        linea.setFrameShape(QFrame.Shape.HLine)
        linea.setFrameShadow(QFrame.Shadow.Sunken)
        linea.setStyleSheet("background-color: #bdc3c7;")
        layout.addWidget(linea)
        
        # Estadísticas en grid
        grid_layout = QHBoxLayout()
        
        # Total guías
        total_widget = QWidget()
        total_layout = QVBoxLayout(total_widget)
        total_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        total_num = QLabel(str(total_guias))
        total_num.setProperty("class", "numero")
        total_num.setStyleSheet("color: #3498db;")
        total_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        total_layout.addWidget(total_num)
        
        total_label = QLabel("TOTAL GUÍAS")
        total_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        total_layout.addWidget(total_label)
        
        grid_layout.addWidget(total_widget)
        
        # Desviadas
        desviadas_widget = QWidget()
        desviadas_layout = QVBoxLayout(desviadas_widget)
        desviadas_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        desviadas_num = QLabel(str(desviadas))
        desviadas_num.setProperty("class", "numero")
        desviadas_num.setStyleSheet("color: #27ae60;")
        desviadas_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desviadas_layout.addWidget(desviadas_num)
        
        desviadas_label = QLabel("DESVIACIONES")
        desviadas_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        desviadas_layout.addWidget(desviadas_label)
        
        grid_layout.addWidget(desviadas_widget)
        
        # Entregadas (ENT)
        entregadas_widget = QWidget()
        entregadas_layout = QVBoxLayout(entregadas_widget)
        entregadas_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        entregadas_num = QLabel(str(entregadas))
        entregadas_num.setProperty("class", "numero")
        entregadas_num.setStyleSheet("color: #f39c12;")
        entregadas_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        entregadas_layout.addWidget(entregadas_num)
        
        entregadas_label = QLabel("ENTREGADAS (ENT)")
        entregadas_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        entregadas_layout.addWidget(entregadas_label)
        
        grid_layout.addWidget(entregadas_widget)
        
        # Errores
        errores_widget = QWidget()
        errores_layout = QVBoxLayout(errores_widget)
        errores_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        errores_num = QLabel(str(errores))
        errores_num.setProperty("class", "numero")
        errores_num.setStyleSheet("color: #e74c3c;")
        errores_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        errores_layout.addWidget(errores_num)
        
        errores_label = QLabel("ERRORES")
        errores_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        errores_layout.addWidget(errores_label)
        
        grid_layout.addWidget(errores_widget)
        
        # Advertencias
        advertencias_widget = QWidget()
        advertencias_layout = QVBoxLayout(advertencias_widget)
        advertencias_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        advertencias_num = QLabel(str(advertencias))
        advertencias_num.setProperty("class", "numero")
        advertencias_num.setStyleSheet("color: #f39c12;")
        advertencias_num.setAlignment(Qt.AlignmentFlag.AlignCenter)
        advertencias_layout.addWidget(advertencias_num)
        
        advertencias_label = QLabel("ADVERTENCIAS")
        advertencias_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        advertencias_layout.addWidget(advertencias_label)
        
        grid_layout.addWidget(advertencias_widget)
        
        layout.addLayout(grid_layout)
        
        # Tiempo total
        tiempo_label = QLabel(f"⏱️ TIEMPO TOTAL: {tiempo_total}")
        tiempo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        tiempo_label.setStyleSheet("font-size: 14pt; font-weight: bold; color: #34495e; margin-top: 15px;")
        layout.addWidget(tiempo_label)
        
        # Botón cerrar
        btn_cerrar = QPushButton("ACEPTAR")
        btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 15px;
                border-radius: 8px;
                font-size: 12pt;
                min-width: 200px;
                margin-top: 20px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        btn_cerrar.clicked.connect(self.accept)
        layout.addWidget(btn_cerrar)

# =======================
# VENTANA DE HISTORIAL
# =======================
class HistorialWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 Historial de Guías Procesadas")
        self.setMinimumSize(900, 500)
        self.setModal(True)
        
        layout = QVBoxLayout(self)
        
        # Título
        titulo = QLabel("📊 GUÍAS PROCESADAS")
        titulo.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(titulo)
        
        # Tabla con 5 columnas
        self.tabla = QTableWidget()
        self.tabla.setColumnCount(5)
        self.tabla.setHorizontalHeaderLabels(["Guía", "Estado", "Resultado", "Navegador", "Fecha/Hora"])
        self.tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tabla.setAlternatingRowColors(True)
        layout.addWidget(self.tabla)
        
        # Botón cerrar
        btn_cerrar = QPushButton("CERRAR")
        btn_cerrar.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                min-width: 100px;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        btn_cerrar.clicked.connect(self.accept)
        layout.addWidget(btn_cerrar)
    
    def actualizar_historial(self, datos):
        """Actualiza la tabla con los datos del historial"""
        self.tabla.setRowCount(len(datos))
        for i, (guia, estado, resultado, nav, fecha) in enumerate(datos):
            self.tabla.setItem(i, 0, QTableWidgetItem(guia))
            
            item_estado = QTableWidgetItem(estado)
            if "✅" in estado:
                item_estado.setForeground(QColor("#27ae60"))
            elif "📦" in estado:
                item_estado.setForeground(QColor("#f39c12"))
            elif "❌" in estado:
                item_estado.setForeground(QColor("#e74c3c"))
            elif "⚠️" in estado:
                item_estado.setForeground(QColor("#f39c12"))
            self.tabla.setItem(i, 1, item_estado)
            
            item_resultado = QTableWidgetItem(resultado)
            if "ENT" in resultado:
                item_resultado.setForeground(QColor("#f39c12"))
                item_resultado.setBackground(QColor("#fff3cd"))
            elif "ADVERTENCIA" in resultado:
                item_resultado.setForeground(QColor("#f39c12"))
                item_resultado.setBackground(QColor("#fff3cd"))
            elif "ERROR" in resultado:
                item_resultado.setForeground(QColor("#e74c3c"))
            self.tabla.setItem(i, 2, item_resultado)
            
            self.tabla.setItem(i, 3, QTableWidgetItem(nav))
            self.tabla.setItem(i, 4, QTableWidgetItem(fecha))

# =======================
# VENTANA DE LOGIN
# =======================
class LoginWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🔐 Iniciar Sesión - ALERTRAN")
        self.setMinimumWidth(200)
        self.setModal(True)
        
        self.setStyleSheet("""
            QDialog {
                rgba(255, 255, 255, 0.95);
                border-radius: 10px;
            }
            QLineEdit {
                padding: 8px;
                border: 2px solid #bdc3c7;
                border-radius: 5px;
                font-size: 11pt;
            }
            QLineEdit:focus {
                border-color: #3498db;
            }
            QLabel {
                font-size: 11pt;
            }
        """)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 30, 30, 30)
        
        # Título
        titulo = QLabel("🔐 INICIAR SESIÓN")
        titulo.setFont(QFont("Arial", 14, QFont.Weight.Bold))
        titulo.setAlignment(Qt.AlignmentFlag.AlignCenter)
        titulo.setStyleSheet("color: #2c3e50; margin-bottom: 10px;")
        layout.addWidget(titulo)
        
        # Formulario
        form_widget = QWidget()
        form_layout = QFormLayout(form_widget)
        form_layout.setSpacing(15)
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        
        self.usuario_input = QLineEdit()
        self.usuario_input.setPlaceholderText("Ingrese su usuario")
        self.usuario_input.setMinimumHeight(35)
        form_layout.addRow("👤 Usuario:", self.usuario_input)
        
        self.password_input = QLineEdit()
        self.password_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_input.setPlaceholderText("Ingrese su contraseña")
        self.password_input.setMinimumHeight(35)
        form_layout.addRow("🔒 Contraseña:", self.password_input)
        
        layout.addWidget(form_widget)
        
        # Botones
        button_layout = QHBoxLayout()
        button_layout.setSpacing(15)
        
        self.btn_login = QPushButton("✅ INICIAR SESIÓN")
        self.btn_login.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                border: 1px solid #bdc3c7;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        self.btn_login.clicked.connect(self.accept)
        
        self.btn_cancel = QPushButton("❌ CANCELAR")
        self.btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: #ebc7c7;
                color: #616060;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                border: 1px solid #bdc3c7;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton:hover {
                background-color: #e84646;
            }
        """)
        self.btn_cancel.clicked.connect(self.reject)
        
        button_layout.addWidget(self.btn_login)
        button_layout.addWidget(self.btn_cancel)
        
        layout.addLayout(button_layout)
        
        # Conectar tecla Enter
        self.usuario_input.returnPressed.connect(self.password_input.setFocus)
        self.password_input.returnPressed.connect(self.accept)
    
    def get_credentials(self):
        return self.usuario_input.text(), self.password_input.text()

# =======================
# SEÑALES PARA COMUNICACIÓN ENTRE HILOS
# =======================
class ProcesoSenales(QObject):
    progreso = Signal(int)
    estado = Signal(str)
    log = Signal(str)
    error = Signal(str)
    finalizado = Signal()
    archivo_errores = Signal(str)
    guia_procesada = Signal(str, str, str, str, str)  # guia, estado, resultado, navegador, fecha
    proceso_cancelado = Signal()
    tiempo_restante = Signal(str)  # Nueva señal para tiempo restante

# =======================
# HILO PARA EL PROCESO PRINCIPAL (MULTI-NAVEGADOR)
# =======================
class ProcesoThread(QThread):
    def __init__(self, usuario, password, ciudad, tipo, ampliacion, excel_path, num_navegadores):
        super().__init__()
        self.usuario = usuario
        self.password = password
        self.ciudad = ciudad
        self.tipo = tipo
        self.ampliacion = ampliacion
        self.excel_path = excel_path
        self.num_navegadores = min(num_navegadores, 6)
        self.senales = ProcesoSenales()
        self.guias_error = []
        self.guias_advertencia = []
        self.guias_ent = []
        self.guias_procesadas_exito = set()
        self.guias_procesadas_ent = set()
        self.guias_en_error = set()
        self.pages = []
        self.browsers = []
        self.contexts = []
        self.lock = asyncio.Lock()
        self.guias_procesadas = set()
        self.cola_guias = []
        self.procesando = True
        self.cancelado = False
        self.tiempo_inicio = None
        self.total_guias = 0
        self.carpeta_descargas = obtener_carpeta_descargas()

    def leer_excel(self, ruta: Union[str, Path]) -> List[str]:
        """Lee el archivo Excel y extrae las guías"""
        wb = load_workbook(ruta, read_only=True, data_only=True)
        try:
            ws = wb.active
            guias = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    guia = str(row[0]).strip()
                    if guia:
                        guias.append(guia)
            return guias
        finally:
            wb.close()

    async def esperar_overlay(self, page, timeout=10000):
        try:
            await page.wait_for_selector("#capa_selector", state="hidden", timeout=timeout)
        except:
            pass
        await asyncio.sleep(1.5)

    async def detectar_error_guia(self, page):
        errores = ["No se encontraron", "Error", "No existe", "sin resultados"]
        for texto in errores:
            try:
                if await page.get_by_text(texto, exact=False).count() > 0:
                    return True
            except:
                pass
        return False

    async def verificar_pagina_activa(self, page):
        try:
            titulo = await page.title()
            return True
        except:
            return False

    async def verificar_estado_ent(self, page, nav_idx):
        """Verifica si la guía tiene estado ENT usando el localizador exacto"""
        try:
            # Usar el localizador exacto
            elemento_ent = (
                page
                .frame_locator("frame[name=\"menu\"]")
                .frame_locator("iframe[name=\"principal\"]")
                .frame_locator("frame[name=\"resultado\"]")
                .get_by_role("cell", name="ENT", exact=True)
            )
            
            if await elemento_ent.count() > 0:
                self.senales.log.emit(f"📦 [Nav{nav_idx}] Estado ENT detectado")
                return True
            return False
        except:
            return False

    async def calcular_tiempo_restante(self, procesadas, total):
        """Calcula y emite el tiempo restante estimado"""
        if self.tiempo_inicio and procesadas > 0:
            elapsed = time.time() - self.tiempo_inicio
            velocidad = procesadas / elapsed if elapsed > 0 else 0
            if velocidad > 0:
                restantes = total - procesadas
                segundos_restantes = restantes / velocidad
                tiempo_restante = str(timedelta(seconds=int(segundos_restantes)))
                self.senales.tiempo_restante.emit(f"⏱️ Tiempo restante: {tiempo_restante}")

    async def hacer_login(self, page, nav_idx):
        try:
            self.senales.log.emit(f"🔐 [Nav{nav_idx}] Iniciando sesión...")
            await page.fill('input[name="j_username"]', self.usuario)
            await asyncio.sleep(0.2)
            await page.fill('input[name="j_password"]', self.password)
            await asyncio.sleep(0.2)
            await page.get_by_role("button", name="Aceptar").click()
            await page.wait_for_load_state("networkidle")
            await asyncio.sleep(1)
            return True
        except Exception as e:
            self.senales.log.emit(f"❌ [Nav{nav_idx}] Error login: {str(e)}")
            return False

    async def navegar_a_funcionalidad_7_8(self, page, nav_idx):
        try:
            self.senales.log.emit(f"🧭 [Nav{nav_idx}] Navegando a 7.8...")
            
            if not await self.verificar_pagina_activa(page):
                return False
            
            menu = page.frame_locator('frame[name="menu"]')
            
            try:
                base_selector = menu.get_by_role("cell", name="ABA BARRANQUILLA AEROPUE").locator("span")
                if await base_selector.count() > 0:
                    await base_selector.click(timeout=3000)
            except:
                pass

            try:
                ciudad_selector = menu.get_by_role("list").get_by_text(self.ciudad)
                if await ciudad_selector.count() > 0:
                    await ciudad_selector.click(timeout=2000)
                    await asyncio.sleep(TIEMPO_ESPERA_CLICK * 2 / 1000)
            except:
                pass

            funcionalidad = menu.locator('input[name="funcionalidad_codigo"]:not([type="hidden"])')
            await funcionalidad.wait_for(state="visible", timeout=20000)
            await funcionalidad.fill("")
            await asyncio.sleep(0.2)
            await funcionalidad.fill("7.8")
            await asyncio.sleep(0.2)
            await funcionalidad.press("Enter")

            await self.esperar_overlay(page)
            await asyncio.sleep(TIEMPO_ESPERA_NAVEGACION / 1000)
            
            self.senales.log.emit(f"✅ [Nav{nav_idx}] Navegación completada")
            return True
            
        except Exception as e:
            self.senales.log.emit(f"❌ [Nav{nav_idx}] Error navegación: {str(e)}")
            return False

    async def ingresar_codigos_con_manejo(self, contenido, tipo, origen, nav_idx):
        try:
            tipo_input = contenido.locator('input[name="tipo_incidencia_codigo"]:not([type="hidden"])')
            await tipo_input.wait_for(state="visible", timeout=10000)
            await tipo_input.fill("")
            await tipo_input.fill(tipo)
            await tipo_input.press("Enter")

            origen_input = contenido.locator('input[name="tipo_origen_incidencia_codigo"]:not([type="hidden"])')
            await origen_input.wait_for(state="visible", timeout=10000)
            await origen_input.fill("")
            await asyncio.sleep(TIEMPO_ESPERA_INGRESO_CODIGOS / 1000)
            await origen_input.fill(origen)
            await asyncio.sleep(TIEMPO_ESPERA_INGRESO_CODIGOS / 1000)
            await origen_input.press("Enter")
            await asyncio.sleep(TIEMPO_ESPERA_CLICK / 1000)

            return True
        except Exception as e:
            self.senales.log.emit(f"⚠️ [Nav{nav_idx}] Error códigos: {str(e)}")
            return False

    async def manejar_boton_volver(self, solapas, guia, nav_idx):
        try:
            self.senales.log.emit(f"⏎ [Nav{nav_idx}] Clic en Volver...")
            await asyncio.sleep(2)
            
            boton_volver = solapas.get_by_role("button", name="Volver")
            
            if await boton_volver.count() == 0:
                self.senales.log.emit(f"⚠️ [Nav{nav_idx}] Botón Volver no encontrado")
                return False
            
            await boton_volver.click(timeout=TIEMPO_ESPERA_VOLVER)
            await self.esperar_overlay(self.pages[nav_idx - 1])
            await asyncio.sleep(3)
            
            return await self.verificar_pagina_activa(self.pages[nav_idx - 1])
            
        except Exception as e:
            self.senales.log.emit(f"⚠️ [Nav{nav_idx}] Error Volver: {str(e)}")
            return False

    async def verificar_incidencia_creada(self, page, nav_idx, guia):
        """Verifica si la incidencia se creó correctamente después de un timeout"""
        try:
            await asyncio.sleep(3)
            
            # Buscar mensajes de éxito en la página
            mensajes_exito = [
                "Incidencia creada", 
                "Éxito", 
                "Success", 
                "Creado correctamente",
                "Operación exitosa",
                "Se ha creado la incidencia"
            ]
            
            for mensaje in mensajes_exito:
                if await page.get_by_text(mensaje, exact=False).count() > 0:
                    self.senales.log.emit(f"✅ [Nav{nav_idx}] Incidencia creada exitosamente (verificado)")
                    return True
            
            # Buscar mensajes de error
            mensajes_error = [
                "Error", 
                "Falló", 
                "No se pudo crear", 
                "Exception",
                "No fue posible",
                "Reintente"
            ]
            
            for mensaje in mensajes_error:
                if await page.get_by_text(mensaje, exact=False).count() > 0:
                    self.senales.log.emit(f"❌ [Nav{nav_idx}] Se detectó mensaje de error: {mensaje}")
                    return False
            
            # Verificar si la guía aparece en el resultado (puede indicar que no se creó)
            try:
                resultado_frame = (
                    page
                    .frame_locator("frame[name=\"menu\"]")
                    .frame_locator("iframe[name=\"principal\"]")
                    .frame_locator("frame[name=\"resultado\"]")
                )
                
                if await resultado_frame.get_by_text(guia, exact=False).count() > 0:
                    # Si la guía sigue apareciendo, podría no haberse creado la incidencia
                    self.senales.log.emit(f"⚠️ [Nav{nav_idx}] La guía aún aparece en resultados - posible no creación")
                    return False
            except:
                pass
            
            # Si no hay mensajes claros, no podemos confirmar
            self.senales.log.emit(f"⚠️ [Nav{nav_idx}] No se pudo confirmar creación - MARCANDO COMO ADVERTENCIA")
            return None  # None indica indeterminado
            
        except Exception as e:
            self.senales.log.emit(f"⚠️ [Nav{nav_idx}] Error en verificación: {str(e)}")
            return None

    async def crear_incidencia(self, page, guia, nav_idx, intento=1):
        try:
            # Verificar si la guía ya fue procesada con éxito
            async with self.lock:
                if guia in self.guias_procesadas_exito:
                    self.senales.log.emit(f"⏭️ [Nav{nav_idx}] Guía {guia} ya procesada con éxito - omitiendo")
                    return True
                if guia in self.guias_procesadas_ent:
                    self.senales.log.emit(f"⏭️ [Nav{nav_idx}] Guía {guia} ya procesada como ENT - omitiendo")
                    return True
                if guia in self.guias_en_error:
                    self.senales.log.emit(f"⏭️ [Nav{nav_idx}] Guía {guia} ya en error - omitiendo")
                    return False
            
            if not await self.verificar_pagina_activa(page):
                raise Exception("Página no activa")
            
            principal = (
                page
                .frame_locator('frame[name="menu"]')
                .frame_locator('iframe[name="principal"]')
            )

            filtro = principal.frame_locator('frame[name="filtro"]')
            resultado = principal.frame_locator('frame[name="resultado"]')
            contenido = principal.frame_locator('frame[name="contenido"]')
            solapas = principal.frame_locator('frame[name="solapas"]')

            # Buscar guía
            try:
                envio = filtro.locator('input[name="nenvio"]:not([type="hidden"])')
                await envio.wait_for(state="visible", timeout=15000)
                await envio.fill("")
                await asyncio.sleep(0.5)
                await envio.fill(guia)
                await asyncio.sleep(0.5)
                await envio.press("Enter")
            except Exception as e:
                error_msg = f"Campo búsqueda no disponible: {str(e)}"
                self.senales.log.emit(f"❌ [Nav{nav_idx}] {error_msg}")
                async with self.lock:
                    self.guias_error.append((guia, error_msg))
                    self.guias_en_error.add(guia)
                fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.senales.guia_procesada.emit(guia, "❌ ERROR", "ERROR BÚSQUEDA", f"Nav{nav_idx}", fecha)
                raise Exception(error_msg)

            await self.esperar_overlay(page)
            await asyncio.sleep(TIEMPO_ESPERA_CLICK / 1000)

            # Verificar si es ENT
            es_ent = await self.verificar_estado_ent(page, nav_idx)
            if es_ent:
                mensaje = f"📦 [Nav{nav_idx}] {guia} - GUÍA ENTREGADA (ENT)"
                self.senales.log.emit(mensaje)
                async with self.lock:
                    self.guias_ent.append(guia)
                    self.guias_procesadas_ent.add(guia)
                
                # Registrar en historial
                fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.senales.guia_procesada.emit(guia, "📦 ENTREGADA", "ENT", f"Nav{nav_idx}", fecha)
                
                # Hacer clic en Volver para limpiar
                try:
                    boton_volver = solapas.get_by_role("button", name="Volver")
                    if await boton_volver.count() > 0:
                        await boton_volver.click(timeout=10000)
                        await self.esperar_overlay(page)
                        await asyncio.sleep(2)
                except:
                    pass
                
                return True

            if await self.detectar_error_guia(page):
                error_msg = "Guía sin resultados"
                self.senales.log.emit(f"❌ [Nav{nav_idx}] {error_msg}")
                async with self.lock:
                    self.guias_error.append((guia, error_msg))
                    self.guias_en_error.add(guia)
                fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.senales.guia_procesada.emit(guia, "❌ SIN RESULTADOS", "SIN RESULTADOS", f"Nav{nav_idx}", fecha)
                raise Exception(error_msg)

            # Abrir guía
            try:
                await resultado.get_by_role("link", name=guia).click(timeout=10000)
            except Exception as e:
                error_msg = f"No se pudo abrir la guía: {str(e)}"
                self.senales.log.emit(f"❌ [Nav{nav_idx}] {error_msg}")
                async with self.lock:
                    self.guias_error.append((guia, error_msg))
                    self.guias_en_error.add(guia)
                fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.senales.guia_procesada.emit(guia, "❌ ERROR APERTURA", "ERROR APERTURA", f"Nav{nav_idx}", fecha)
                raise Exception(error_msg)

            await self.esperar_overlay(page)
            await asyncio.sleep(TIEMPO_ESPERA_CLICK / 1000)

            # Ingresar códigos
            exito_codigos = await self.ingresar_codigos_con_manejo(contenido, self.tipo, "018", nav_idx)
            if not exito_codigos:
                error_msg = "Error ingresando códigos"
                self.senales.log.emit(f"❌ [Nav{nav_idx}] {error_msg}")
                async with self.lock:
                    self.guias_error.append((guia, error_msg))
                    self.guias_en_error.add(guia)
                fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                self.senales.guia_procesada.emit(guia, "❌ ERROR CÓDIGOS", "ERROR CÓDIGOS", f"Nav{nav_idx}", fecha)
                raise Exception(error_msg)

            # Ingresar ampliación
            await contenido.locator('textarea[name="ampliacion_incidencia"]').fill(self.ampliacion)

            # Crear incidencia con manejo mejorado de timeout y verificación
            incidencia_creada = False
            incidencia_indeterminada = False
            
            try:
                async with page.expect_popup(timeout=10000) as pop_info:
                    await contenido.get_by_role("button", name="Crear").click()
                popup = await pop_info.value
                await popup.close()
                await asyncio.sleep(2)
                incidencia_creada = True
                self.senales.log.emit(f"✅ [Nav{nav_idx}] Popup cerrado correctamente")
            except Exception as e:
                self.senales.log.emit(f"⚠️ [Nav{nav_idx}] Timeout/Error en creación - {str(e)}")
                
                # Verificar si la incidencia se creó a pesar del timeout
                self.senales.log.emit(f"🔍 [Nav{nav_idx}] Verificando si la incidencia se creó correctamente...")
                resultado_verificacion = await self.verificar_incidencia_creada(page, nav_idx, guia)
                
                if resultado_verificacion is True:
                    incidencia_creada = True
                elif resultado_verificacion is False:
                    incidencia_creada = False
                else:
                    # Indeterminado - lo tratamos como advertencia
                    incidencia_indeterminada = True
                    self.senales.log.emit(f"⚠️ [Nav{nav_idx}] ESTADO INDETERMINADO - Se marcará como advertencia")

            # Volver
            exito_volver = await self.manejar_boton_volver(solapas, guia, nav_idx)
            
            # Registrar según el resultado
            fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            if incidencia_creada and exito_volver:
                # Éxito completo
                async with self.lock:
                    self.guias_procesadas_exito.add(guia)
                self.senales.guia_procesada.emit(guia, "✅ PROCESADA", "COMPLETADO", f"Nav{nav_idx}", fecha)
                self.senales.log.emit(f"✅ [Nav{nav_idx}] {guia} OK")
            elif incidencia_indeterminada:
                # Indeterminado - lo tratamos como advertencia para evitar doble procesamiento
                error_msg = "Estado indeterminado - posible creación no confirmada"
                async with self.lock:
                    self.guias_advertencia.append((guia, f"[Nav{nav_idx}] {error_msg}"))
                    # No lo marcamos como error para no reprocesar, pero lo registramos
                self.senales.guia_procesada.emit(guia, "⚠️ ADVERTENCIA", "NO CONFIRMADO", f"Nav{nav_idx}", fecha)
                self.senales.log.emit(f"⚠️ [Nav{nav_idx}] {guia} - ADVERTENCIA: {error_msg}")
            else:
                # Error
                error_msg = "Error en procesamiento"
                if not incidencia_creada:
                    error_msg = "Incidencia no creada"
                if not exito_volver:
                    error_msg = "Error al volver"
                
                async with self.lock:
                    self.guias_error.append((guia, f"[Nav{nav_idx}] {error_msg}"))
                    self.guias_en_error.add(guia)
                self.senales.guia_procesada.emit(guia, "❌ ERROR", error_msg, f"Nav{nav_idx}", fecha)
                self.senales.log.emit(f"❌ [Nav{nav_idx}] {guia}: {error_msg}")
            
            if not exito_volver and intento < MAX_REINTENTOS and not incidencia_creada:
                self.senales.log.emit(f"🔄 [Nav{nav_idx}] Reintento {intento+1}/{MAX_REINTENTOS}")
                return await self.crear_incidencia(page, guia, nav_idx, intento + 1)
            
            return exito_volver or incidencia_creada

        except Exception as e:
            error_msg = str(e)
            self.senales.log.emit(f"❌ [Nav{nav_idx}] Error fatal en creación: {error_msg}")
            
            # Registrar error si no se hizo ya
            async with self.lock:
                if guia not in self.guias_en_error and guia not in self.guias_procesadas_exito:
                    self.guias_error.append((guia, f"[Nav{nav_idx}] {error_msg}"))
                    self.guias_en_error.add(guia)
                    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    self.senales.guia_procesada.emit(guia, "❌ ERROR", error_msg, f"Nav{nav_idx}", fecha)
            
            if intento < MAX_REINTENTOS and "Guía sin resultados" not in error_msg:
                await asyncio.sleep(3)
                return await self.crear_incidencia(page, guia, nav_idx, intento + 1)
            return False

    async def trabajador_navegador(self, nav_idx, total_guias, resultados):
        """Worker para cada navegador"""
        try:
            page = self.pages[nav_idx - 1]
            guias_procesadas_local = 0
            
            while self.procesando and not self.cancelado:
                # Obtener siguiente guía de la cola
                async with self.lock:
                    if not self.cola_guias:
                        break
                    guia = self.cola_guias.pop(0)
                    
                    # Verificar si la guía ya fue procesada por otro navegador
                    if (guia in self.guias_procesadas_exito or 
                        guia in self.guias_procesadas_ent or 
                        guia in self.guias_en_error):
                        self.senales.log.emit(f"⏭️ [Nav{nav_idx}] Guía {guia} ya procesada - saltando")
                        continue
                
                try:
                    self.senales.log.emit(f"🌐 [Nav{nav_idx}] Procesando: {guia}")
                    
                    exito = await self.crear_incidencia(page, guia, nav_idx)
                    
                    if exito:
                        guias_procesadas_local += 1
                        async with self.lock:
                            resultados['exitosas'] += 1
                    
                except Exception as e:
                    self.senales.log.emit(f"❌ [Nav{nav_idx}] Error procesando {guia}: {str(e)}")
                
                # Actualizar progreso y tiempo restante
                async with self.lock:
                    resultados['progreso'] += 1
                    progreso = int(resultados['progreso'] / total_guias * 100)
                    self.senales.progreso.emit(progreso)
                    await self.calcular_tiempo_restante(resultados['progreso'], total_guias)
                    self.senales.estado.emit(
                        f"Progreso: {resultados['progreso']}/{total_guias} ({progreso}%) "
                        f"- Éxitos: {resultados['exitosas']} [Navs: {self.num_navegadores}]"
                    )
            
            if self.cancelado:
                self.senales.log.emit(f"🛑 [Nav{nav_idx}] Proceso cancelado por usuario")
            
            self.senales.log.emit(f"📊 [Nav{nav_idx}] Procesó {guias_procesadas_local} guías")
            
        except Exception as e:
            self.senales.log.emit(f"❌ [Nav{nav_idx}] Error fatal: {str(e)}")

    async def proceso_principal(self):
        """Método principal con múltiples navegadores"""
        try:
            # Leer guías
            guias = self.leer_excel(self.excel_path)
            self.total_guias = len(guias)
            
            if self.total_guias == 0:
                self.senales.error.emit("El archivo Excel no contiene guías")
                return

            self.tiempo_inicio = time.time()
            self.senales.estado.emit(f"Procesando {self.total_guias} guías con {self.num_navegadores} navegador(es)...")

            # Iniciar Playwright
            async with async_playwright() as p:
                # Inicializar navegadores
                for i in range(self.num_navegadores):
                    if self.cancelado:
                        break
                        
                    self.senales.log.emit(f"🚀 Iniciando navegador {i+1}/{self.num_navegadores}...")
                    
                    browser = await p.chromium.launch(
                        headless=False,
                        args=['--start-maximized', '--disable-dev-shm-usage']
                    )
                    
                    context = await browser.new_context(
                        viewport={'width': 1280, 'height': 800},
                        locale="es-ES"
                    )
                    
                    page = await context.new_page()
                    page.set_default_timeout(60000)
                    
                    self.browsers.append(browser)
                    self.contexts.append(context)
                    self.pages.append(page)
                    
                    # Login
                    await page.goto("https://alertran.latinlogistics.com.co/padua/inicio.do", timeout=60000)
                    await asyncio.sleep(3)
                    
                    exito_login = await self.hacer_login(page, i+1)
                    if not exito_login:
                        self.senales.error.emit(f"Error login navegador {i+1}")
                        return
                    
                    # Navegar a 7.8
                    exito_navegacion = await self.navegar_a_funcionalidad_7_8(page, i+1)
                    if not exito_navegacion:
                        self.senales.error.emit(f"Error navegación navegador {i+1}")
                        return

                if not self.cancelado:
                    # Crear cola de guías
                    self.cola_guias = guias.copy()
                    
                    # Resultados compartidos
                    resultados = {'progreso': 0, 'exitosas': 0}

                    # Iniciar workers
                    tareas = []
                    for i in range(self.num_navegadores):
                        tarea = self.trabajador_navegador(i+1, self.total_guias, resultados)
                        tareas.append(tarea)

                    await asyncio.gather(*tareas)

                # Cerrar navegadores
                for browser in self.browsers:
                    await browser.close()

                if not self.cancelado:
                    # Guardar errores y advertencias
                    if self.guias_error or self.guias_advertencia:
                        ruta = self.guardar_excel_errores()
                        self.senales.archivo_errores.emit(ruta)
                        self.senales.log.emit(f"\n📊 Errores: {len(self.guias_error)}")
                        self.senales.log.emit(f"⚠️ Advertencias: {len(self.guias_advertencia)}")
                    
                    # Mostrar resumen de guías ENT
                    if self.guias_ent:
                        self.senales.log.emit(f"\n📦 GUÍAS ENTREGADAS (ENT): {len(self.guias_ent)}")
                        for guia in self.guias_ent[:10]:
                            self.senales.log.emit(f"   • {guia}")
                        if len(self.guias_ent) > 10:
                            self.senales.log.emit(f"   ... y {len(self.guias_ent) - 10} más")
                    else:
                        self.senales.log.emit("\n📦 No se encontraron guías con estado ENT")
                    
                    tiempo_total = time.time() - self.tiempo_inicio
                    tiempo_formateado = str(timedelta(seconds=int(tiempo_total)))
                    
                    self.senales.log.emit(f"\n✨ Completado en {tiempo_formateado}")
                    self.senales.log.emit(f"✨ Desviaciones creadas: {resultados['exitosas'] - len(self.guias_ent)}")
                    self.senales.log.emit(f"✨ Guías ENT (omitidas): {len(self.guias_ent)}")
                    self.senales.log.emit(f"✨ Errores: {len(self.guias_error)}")
                    self.senales.log.emit(f"✨ Advertencias: {len(self.guias_advertencia)}")
                    self.senales.log.emit(f"✨ Total procesado: {resultados['progreso']}/{self.total_guias}")
                    
                    # Emitir señal para mostrar resumen
                    self.senales.finalizado.emit()
                else:
                    self.senales.proceso_cancelado.emit()

        except Exception as e:
            self.senales.error.emit(f"Error: {str(e)}")

    def guardar_excel_errores(self):
        """Guarda el archivo de errores en la carpeta de Descargas"""
        if not self.guias_error and not self.guias_advertencia:
            return None
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores y Advertencias"
        
        # Encabezados
        ws.append(["Guía", "Motivo", "Tipo", "Fecha/Hora"])
        
        # Agregar errores
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for guia, motivo in self.guias_error:
            ws.append([guia, motivo, "ERROR", fecha_actual])
        
        # Agregar advertencias
        for guia, motivo in self.guias_advertencia:
            ws.append([guia, motivo, "ADVERTENCIA", fecha_actual])
        
        # Generar nombre de archivo único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_nombre = f"errores_alertran_{timestamp}"
        
        # Guardar en carpeta de descargas
        ruta_archivo = self.carpeta_descargas / f"{base_nombre}.xlsx"
        
        # Asegurar nombre único
        contador = 1
        while ruta_archivo.exists():
            ruta_archivo = self.carpeta_descargas / f"{base_nombre}_{contador}.xlsx"
            contador += 1
        
        wb.save(ruta_archivo)
        return str(ruta_archivo)

    def guardar_log_completo(self, log_contenido):
        """Guarda el log completo en la carpeta de Descargas"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base_nombre = f"log_alertran_{timestamp}"
        
        # Guardar en carpeta de descargas
        ruta_archivo = self.carpeta_descargas / f"{base_nombre}.txt"
        
        # Asegurar nombre único
        contador = 1
        while ruta_archivo.exists():
            ruta_archivo = self.carpeta_descargas / f"{base_nombre}_{contador}.txt"
            contador += 1
        
        with open(ruta_archivo, 'w', encoding='utf-8') as f:
            f.write(log_contenido)
        
        return str(ruta_archivo)

    def cancelar(self):
        """Cancela el proceso"""
        self.cancelado = True
        self.procesando = False

    def run(self):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(self.proceso_principal())
        finally:
            loop.close()

# =======================
# VENTANA PRINCIPAL
# =======================
class VentanaPrincipal(QMainWindow):
    def __init__(self):
        super().__init__()
        self.excel_path = None
        self.proceso_thread = None
        self.sesion_activa = False
        self.usuario_actual = ""
        self.password_actual = ""
        self.historial_datos = []  # Lista para guardar el historial (guia, estado, resultado, navegador, fecha)
        self.historial_window = None
        self.tiempo_inicio = None
        self.total_guias = 0
        self.guias_ent = []
        self.guias_error_count = 0
        self.guias_advertencia_count = 0
        self.desviaciones_creadas = 0
        self.carpeta_descargas = obtener_carpeta_descargas()
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("ALERTRAN - Gestión Desviaciones")
        self.setMinimumSize(800, 800)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        layout_principal = QVBoxLayout(central_widget)
        layout_principal.setSpacing(10)
        layout_principal.setContentsMargins(20, 20, 20, 20)

        # ===== PANEL DE SESIÓN =====
        grupo_sesion = QGroupBox("🔐 CONTROL DE ACCESO")
        grupo_sesion.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11pt; }")
        
        layout_sesion = QHBoxLayout(grupo_sesion)
        layout_sesion.setSpacing(5)
        
        self.btn_login = QPushButton("🔑 INICIAR SESIÓN")
        self.btn_login.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.btn_login.clicked.connect(self.abrir_login)
        layout_sesion.addWidget(self.btn_login)
        
        self.btn_logout = QPushButton("🚪 CERRAR SESIÓN")
        self.btn_logout.setStyleSheet("""
            QPushButton {
                background-color: #ebc7c7;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                border: 1px solid #ced6db;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #d35400; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.btn_logout.clicked.connect(self.cerrar_sesion)
        self.btn_logout.setEnabled(False)
        layout_sesion.addWidget(self.btn_logout)
        
        self.lbl_estado_sesion = QLabel("⛔ SESIÓN NO INICIADA")
        self.lbl_estado_sesion.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_estado_sesion.setStyleSheet("""
            QLabel {
                background-color: #fdeded;
                color: #e74c3c;
                font-weight: bold;
                padding: 8px;
                border-radius: 5px;
                border: 1px solid #e74c3c;
                min-width: 150px;
            }
        """)
        layout_sesion.addWidget(self.lbl_estado_sesion)
        
        layout_principal.addWidget(grupo_sesion)

        # ===== CONFIGURACIÓN =====
        grupo_config = QGroupBox("⚙️ CONFIGURACIÓN")
        grupo_config.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11pt; }")
        
        layout_config = QFormLayout(grupo_config)
        layout_config.setSpacing(5)

        self.ciudad_combo = QComboBox()
        self.ciudad_combo.addItems(CIUDADES)
        self.ciudad_combo.setCurrentText("ABA")
        self.ciudad_combo.setMinimumHeight(20)
        layout_config.addRow("📍 Regional :", self.ciudad_combo)
        
        self.tipo_combo = QComboBox()
        self.tipo_combo.addItems(TIPOS_INCIDENCIA)
        self.tipo_combo.setCurrentText("22")
        self.tipo_combo.setMinimumHeight(20)
        layout_config.addRow("📌 desviación :", self.tipo_combo)
        
        self.ampliacion_input = QLineEdit()
        self.ampliacion_input.setPlaceholderText("Amplación Desviación :")
        self.ampliacion_input.setMinimumHeight(20)
        layout_config.addRow("📝 Ampliación:", self.ampliacion_input)
        
        # Selector de navegadores
        nav_layout = QHBoxLayout()
        self.num_navegadores_spin = QSpinBox()
        self.num_navegadores_spin.setMinimum(1)
        self.num_navegadores_spin.setMaximum(6)
        self.num_navegadores_spin.setValue(1)
        self.num_navegadores_spin.setMinimumHeight(25)
        self.num_navegadores_spin.setPrefix("🚀 ")
        self.num_navegadores_spin.setSuffix(" navegador(es)")
        nav_layout.addWidget(QLabel("Navegadores:"))
        nav_layout.addWidget(self.num_navegadores_spin)
        nav_layout.addStretch()
        layout_config.addRow("", nav_layout)
        
        layout_principal.addWidget(grupo_config)

        # ===== ARCHIVO EXCEL =====
        grupo_excel = QGroupBox("📁 ARCHIVO DE GUÍAS")
        grupo_excel.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11pt; }")
        
        layout_excel = QVBoxLayout(grupo_excel)
        layout_boton_excel = QHBoxLayout()
        
        self.btn_cargar_excel = QPushButton("📂 CARGAR EXCEL")
        self.btn_cargar_excel.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                border: 1px solid #ced6db;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #e67e22; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.btn_cargar_excel.clicked.connect(self.cargar_excel)
        self.btn_cargar_excel.setEnabled(False)
        layout_boton_excel.addWidget(self.btn_cargar_excel)
        
        self.lbl_archivo = QLabel("❌ NINGÚN ARCHIVO")
        self.lbl_archivo.setStyleSheet("color: #e74c3c; font-style: italic;")
        layout_boton_excel.addWidget(self.lbl_archivo)
        layout_boton_excel.addStretch()
        
        layout_excel.addLayout(layout_boton_excel)
        layout_principal.addWidget(grupo_excel)

        # ===== PROGRESO =====
        grupo_progreso = QGroupBox("📊 PROGRESO")
        grupo_progreso.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11pt; }")
        
        layout_progreso = QVBoxLayout(grupo_progreso)
        
        # Barra de progreso estilo Mac
        self.progress_bar = MacProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout_progreso.addWidget(self.progress_bar)
        
        # Label para tiempo restante
        self.lbl_tiempo_restante = QLabel("")
        self.lbl_tiempo_restante.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_tiempo_restante.setStyleSheet("font-size: 10pt; color: #3498db; font-weight: bold; padding: 5px;")
        layout_progreso.addWidget(self.lbl_tiempo_restante)
        
        self.lbl_estado = QLabel("💤 LISTO")
        self.lbl_estado.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.lbl_estado.setStyleSheet("font-weight: bold; font-size: 11pt;")
        layout_progreso.addWidget(self.lbl_estado)
        
        layout_principal.addWidget(grupo_progreso)

        # ===== LOG =====
        grupo_log = QGroupBox("📋 REGISTRO DE ACTIVIDAD")
        grupo_log.setStyleSheet("QGroupBox { font-weight: bold; font-size: 11pt; }")
        
        layout_log = QVBoxLayout(grupo_log)
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setMinimumHeight(150)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background-color: #2c3e50;
                font-family: 'Consolas', monospace;
                font-size: 10pt;
                border: 2px solid #34495e;
                border-radius: 5px;
                color: #ecf0f1;
                padding: 8px;
            }
        """)
        layout_log.addWidget(self.log_text)
        
        layout_principal.addWidget(grupo_log)

        # ===== BOTONES ACCIÓN =====
        layout_botones = QHBoxLayout()
        layout_botones.setSpacing(10)
        
        self.btn_iniciar = QPushButton("▶ INICIAR PROCESO")
        self.btn_iniciar.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                font-size: 12pt;
                min-width: 180px;
            }
            QPushButton:hover { background-color: #2ecc71; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.btn_iniciar.clicked.connect(self.iniciar_proceso)
        self.btn_iniciar.setEnabled(False)
        layout_botones.addWidget(self.btn_iniciar)
        
        self.btn_cancelar = QPushButton("⏹ CANCELAR PROCESO")
        self.btn_cancelar.setStyleSheet("""
            QPushButton {
                background-color: #e67e22;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                font-size: 12pt;
                min-width: 180px;
            }
            QPushButton:hover { background-color: #d35400; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.btn_cancelar.clicked.connect(self.cancelar_proceso)
        self.btn_cancelar.setEnabled(False)
        layout_botones.addWidget(self.btn_cancelar)
        
        self.btn_historial = QPushButton("📋 VER HISTORIAL")
        self.btn_historial.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                font-size: 12pt;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.btn_historial.clicked.connect(self.ver_historial)
        layout_botones.addWidget(self.btn_historial)
        
        self.btn_errores = QPushButton("📥 EXCEL ERRORES")
        self.btn_errores.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #c0392b; }
            QPushButton:disabled { background-color: #95a5a6; }
        """)
        self.btn_errores.clicked.connect(self.mostrar_errores)
        self.btn_errores.setEnabled(False)
        layout_botones.addWidget(self.btn_errores)
        
        self.btn_descargar_log = QPushButton("💾 DESCARGAR LOG")
        self.btn_descargar_log.setStyleSheet("""
            QPushButton {
                background-color: #9b59b6;
                color: white;
                font-weight: bold;
                padding: 12px;
                border-radius: 5px;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton:hover { background-color: #8e44ad; }
        """)
        self.btn_descargar_log.clicked.connect(self.descargar_log)
        layout_botones.addWidget(self.btn_descargar_log)
        
        layout_principal.addLayout(layout_botones)

        # Versión
        info_label = QLabel("🤖 V.7.3")
        info_label.setStyleSheet("color: #3498db; font-size: 9pt; font-weight: bold;")
        info_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout_principal.addWidget(info_label)
        
        self.showMaximized()  # Abrir maximizado

    def abrir_login(self):
        login = LoginWindow(self)
        if login.exec() == QDialog.DialogCode.Accepted:
            usuario, password = login.get_credentials()
            if usuario and password:
                self.usuario_actual = usuario
                self.password_actual = password
                self.sesion_activa = True
                self.actualizar_estado_sesion()
                self.log(f"✅ Sesión iniciada: {usuario}")
                self.habilitar_controles(True)

    def cerrar_sesion(self):
        reply = QMessageBox.question(self, "Cerrar Sesión", 
                                    f"¿Cerrar sesión de {self.usuario_actual}?",
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.sesion_activa = False
            self.usuario_actual = ""
            self.password_actual = ""
            self.actualizar_estado_sesion()
            self.log("🔒 Sesión cerrada")
            self.habilitar_controles(False)
            self.historial_datos.clear()

    def actualizar_estado_sesion(self):
        if self.sesion_activa:
            self.lbl_estado_sesion.setText(f"✅ ACTIVA - {self.usuario_actual}")
            self.lbl_estado_sesion.setStyleSheet("""
                QLabel {
                    background-color: #e8f8f5;
                    color: #27ae60;
                    font-weight: bold;
                    padding: 8px;
                    border-radius: 5px;
                    border: 2px solid #27ae60;
                }
            """)
            self.btn_login.setEnabled(False)
            self.btn_logout.setEnabled(True)
        else:
            self.lbl_estado_sesion.setText("⛔ SESIÓN NO INICIADA")
            self.lbl_estado_sesion.setStyleSheet("""
                QLabel {
                    background-color: #fdeded;
                    color: #e74c3c;
                    font-weight: bold;
                    padding: 8px;
                    border-radius: 5px;
                    border: 2px solid #e74c3c;
                }
            """)
            self.btn_login.setEnabled(True)
            self.btn_logout.setEnabled(False)

    def habilitar_controles(self, habilitar):
        self.btn_cargar_excel.setEnabled(habilitar)
        self.btn_iniciar.setEnabled(habilitar)
        if not habilitar:
            self.excel_path = None
            self.lbl_archivo.setText("❌ NINGÚN ARCHIVO")
            self.lbl_archivo.setStyleSheet("color: #e74c3c; font-style: italic;")
            self.progress_bar.setValue(0)
            self.lbl_estado.setText("💤 LISTO")
            self.lbl_tiempo_restante.setText("")

    def cargar_excel(self):
        archivo, _ = QFileDialog.getOpenFileName(
            self, "Seleccionar Excel", str(Path.home()), "Excel (*.xlsx)"
        )
        if archivo:
            self.excel_path = archivo
            nombre = Path(archivo).name
            
            # Leer el archivo Excel para contar las guías
            try:
                wb = load_workbook(archivo, read_only=True, data_only=True)
                ws = wb.active
                guias_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and str(row[0]).strip():
                        guias_count += 1
                wb.close()
                
                # Actualizar el label con la información
                self.lbl_archivo.setText(f"📄 {nombre} ({guias_count} guías)")
                self.lbl_archivo.setStyleSheet("color: #27ae60; font-weight: bold;")
                self.log(f"✅ Archivo cargado: {nombre} - {guias_count} guías a procesar")
                
                # Guardar el total de guías
                self.total_guias = guias_count
                
            except Exception as e:
                self.lbl_archivo.setText(f"📄 {nombre} (Error al leer)")
                self.lbl_archivo.setStyleSheet("color: #e74c3c; font-weight: bold;")
                self.log(f"⚠️ Error al leer el archivo: {str(e)}")

    def log(self, mensaje):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{ts}] {mensaje}")
        cursor = self.log_text.textCursor()
        cursor.movePosition(QTextCursor.MoveOperation.End)
        self.log_text.setTextCursor(cursor)

    def descargar_log(self):
        """Descarga el contenido del log a la carpeta de Descargas"""
        try:
            # Obtener el contenido del log
            contenido_log = self.log_text.toPlainText()
            
            # Generar nombre de archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_nombre = f"log_alertran_{timestamp}"
            
            # Guardar en carpeta de descargas
            ruta_archivo = self.carpeta_descargas / f"{base_nombre}.txt"
            
            # Asegurar nombre único
            contador = 1
            while ruta_archivo.exists():
                ruta_archivo = self.carpeta_descargas / f"{base_nombre}_{contador}.txt"
                contador += 1
            
            with open(ruta_archivo, 'w', encoding='utf-8') as f:
                f.write(contenido_log)
            
            QMessageBox.information(
                self, "✅ Éxito", 
                f"📄 Log guardado en:\n{ruta_archivo}\n\n"
                f"📁 Carpeta: Descargas"
            )
            
            self.log(f"✅ Log guardado en: {ruta_archivo}")
            
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo guardar el log:\n{str(e)}")

    def ver_historial(self):
        """Muestra la ventana de historial"""
        if not self.historial_window:
            self.historial_window = HistorialWindow(self)
        
        self.historial_window.actualizar_historial(self.historial_datos)
        self.historial_window.show()

    def agregar_al_historial(self, guia, estado, resultado, navegador, fecha):
        """Agrega una guía al historial"""
        self.historial_datos.append((guia, estado, resultado, navegador, fecha))
        # Mantener solo los últimos 1000 registros
        if len(self.historial_datos) > 1000:
            self.historial_datos = self.historial_datos[-1000:]
        
        # Actualizar contadores
        if "📦" in estado:
            self.guias_ent.append(guia)
        elif "❌" in estado:
            self.guias_error_count += 1
        elif "⚠️" in estado:
            self.guias_advertencia_count += 1
        elif "✅" in estado:
            self.desviaciones_creadas += 1

    def actualizar_tiempo_restante(self, tiempo):
        """Actualiza el label de tiempo restante"""
        self.lbl_tiempo_restante.setText(tiempo)

    def mostrar_resumen(self):
        """Muestra la ventana de resumen al finalizar"""
        tiempo_total = datetime.now() - self.tiempo_inicio if self.tiempo_inicio else timedelta(0)
        tiempo_formateado = str(tiempo_total).split('.')[0]  # Quitar microsegundos
        
        resumen = ResumenWindow(
            total_guias=self.total_guias,
            desviadas=self.desviaciones_creadas,
            entregadas=len(self.guias_ent),
            errores=self.guias_error_count,
            advertencias=self.guias_advertencia_count,
            tiempo_total=tiempo_formateado,
            parent=self
        )
        resumen.exec()
        
        # Preguntar si quiere abrir la carpeta de descargas si hay errores
        if self.guias_error_count > 0 or self.guias_advertencia_count > 0:
            reply = QMessageBox.question(
                self, "📂 Abrir Carpeta",
                "¿Desea abrir la carpeta donde se guardaron los archivos de errores?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                if os.name == 'nt':  # Windows
                    os.startfile(self.carpeta_descargas)
                else:  # macOS/Linux
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', str(self.carpeta_descargas)])

    def iniciar_proceso(self):
        if not all([self.sesion_activa, self.ampliacion_input.text(), self.excel_path]):
            QMessageBox.warning(self, "Error", "Complete todos los campos")
            return

        num_nav = self.num_navegadores_spin.value()
        
        reply = QMessageBox.question(
            self, "🔔 CONFIRMAR PROCESO",
            f"📊 **DETALLES DE LA OPERACIÓN**\n\n"
            f"🌐 Navegadores: {num_nav}\n"
            f"👤 Usuario: {self.usuario_actual}\n"
            f"📋 Total guías: {self.total_guias}\n"
            f"📍 Regional: {self.ciudad_combo.currentText()}\n"
            f"📌 Desviación: {self.tipo_combo.currentText()}\n\n"
            f"📁 Los archivos se guardarán en:\n{self.carpeta_descargas}\n\n"
            f"¿Desea continuar?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.No:
            return

        # Resetear contadores
        self.tiempo_inicio = datetime.now()
        self.guias_ent = []
        self.guias_error_count = 0
        self.guias_advertencia_count = 0
        self.desviaciones_creadas = 0

        self.btn_iniciar.setEnabled(False)
        self.btn_cancelar.setEnabled(True)
        self.btn_cargar_excel.setEnabled(False)
        self.btn_errores.setEnabled(False)
        self.btn_login.setEnabled(False)
        self.btn_logout.setEnabled(False)
        self.num_navegadores_spin.setEnabled(False)
        self.progress_bar.setValue(0)
        self.lbl_tiempo_restante.setText("⏱️ Calculando tiempo restante...")
        self.log_text.clear()
        self.historial_datos.clear()

        self.log(f"🚀 Iniciando con {num_nav} navegador(es)...")
        self.log(f"👤 Usuario: {self.usuario_actual}")
        self.log(f"📊 Total guías a procesar: {self.total_guias}")
        self.log(f"📁 Los archivos se guardarán en: {self.carpeta_descargas}")

        self.proceso_thread = ProcesoThread(
            self.usuario_actual,
            self.password_actual,
            self.ciudad_combo.currentText(),
            self.tipo_combo.currentText(),
            self.ampliacion_input.text(),
            self.excel_path,
            num_nav
        )

        self.proceso_thread.senales.progreso.connect(self.progress_bar.setValue)
        self.proceso_thread.senales.estado.connect(self.lbl_estado.setText)
        self.proceso_thread.senales.log.connect(self.log)
        self.proceso_thread.senales.error.connect(self.mostrar_error)
        self.proceso_thread.senales.finalizado.connect(self.proceso_finalizado)
        self.proceso_thread.senales.archivo_errores.connect(self.archivo_errores_generado)
        self.proceso_thread.senales.guia_procesada.connect(self.agregar_al_historial)
        self.proceso_thread.senales.proceso_cancelado.connect(self.proceso_cancelado)
        self.proceso_thread.senales.tiempo_restante.connect(self.actualizar_tiempo_restante)

        self.proceso_thread.start()

    def cancelar_proceso(self):
        """Cancela el proceso en ejecución"""
        reply = QMessageBox.question(
            self, "Cancelar Proceso",
            "¿Está seguro que desea cancelar el proceso?\n\nLas guías no procesadas quedarán pendientes.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes and self.proceso_thread:
            self.log("🛑 Cancelando proceso...")
            self.proceso_thread.cancelar()
            self.btn_cancelar.setEnabled(False)
            self.btn_cancelar.setText("⏹ CANCELANDO...")

    def proceso_cancelado(self):
        """Se llama cuando el proceso se cancela"""
        self.log("✅ Proceso cancelado por usuario")
        self.btn_cancelar.setText("⏹ CANCELAR PROCESO")
        self.lbl_tiempo_restante.setText("")
        self.proceso_finalizado()

    def mostrar_error(self, mensaje):
        QMessageBox.critical(self, "Error", mensaje)
        self.log(f"🔴 ERROR: {mensaje}")
        self.lbl_tiempo_restante.setText("")
        self.proceso_finalizado()

    def proceso_finalizado(self):
        self.btn_iniciar.setEnabled(True)
        self.btn_cancelar.setEnabled(False)
        self.btn_cancelar.setText("⏹ CANCELAR PROCESO")
        self.btn_cargar_excel.setEnabled(True)
        self.btn_login.setEnabled(False)
        self.btn_logout.setEnabled(True)
        self.num_navegadores_spin.setEnabled(True)
        self.lbl_estado.setText("✅ Finalizado")
        
        # Mostrar ventana de resumen
        self.mostrar_resumen()

    def archivo_errores_generado(self, ruta):
        """Se llama cuando se genera el archivo de errores"""
        self.btn_errores.setEnabled(True)
        self.error_path = ruta
        
        # Mostrar mensaje más informativo
        QMessageBox.information(
            self, "✅ Archivo Generado", 
            f"📊 Archivo de errores y advertencias guardado en:\n{ruta}\n\n"
            f"📁 Carpeta: Descargas"
        )
        
        self.log(f"✅ Archivo de errores guardado en: {ruta}")

    def mostrar_errores(self):
        """Muestra la ubicación del archivo de errores y ofrece abrir la carpeta"""
        if hasattr(self, 'error_path') and self.error_path:
            reply = QMessageBox.question(
                self, "📂 Archivo de Errores",
                f"📊 Archivo guardado en:\n{self.error_path}\n\n"
                f"¿Desea abrir la carpeta que contiene el archivo?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                # Abrir la carpeta en el explorador de archivos
                carpeta = Path(self.error_path).parent
                if os.name == 'nt':  # Windows
                    os.startfile(carpeta)
                else:  # macOS/Linux
                    subprocess.run(['open' if sys.platform == 'darwin' else 'xdg-open', str(carpeta)])

# =======================
# MAIN
# =======================
def main():
    app = QApplication(sys.argv)
    loop = qasync.QEventLoop(app)
    asyncio.set_event_loop(loop)
    app.setStyle('Fusion')
    
    ventana = VentanaPrincipal()
    ventana.show()
    
    with loop:
        sys.exit(loop.run_forever())

if __name__ == "__main__":
    main()
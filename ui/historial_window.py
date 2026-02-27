# ui/historial_window.py
"""
Ventana de historial de guías procesadas
"""
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QMessageBox, QApplication
)
from PySide6.QtGui import QFont, QColor
from PySide6.QtCore import Qt, QTimer
from datetime import datetime
from pathlib import Path
import os
import subprocess

from utils.file_utils import FileUtils

class HistorialWindow(QDialog):
    """Ventana para mostrar el historial de guías procesadas"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("📋 Historial de Guías Procesadas")
        self.setMinimumSize(1000, 600)
        self.setModal(True)
        
        self.datos_completos = []
        self.datos_filtrados = []
        self.filtro_actual = "Todos"
        self.carpeta_descargas = FileUtils.obtener_carpeta_descargas()
        self.file_utils = FileUtils()
        
        self._setup_ui()
        self._setup_styles()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Toolbar
        toolbar = self._crear_toolbar()
        layout.addLayout(toolbar)
        
        # Tabla
        self.tabla = self._crear_tabla()
        layout.addWidget(self.tabla)
        
        # Barra de estado
        self.status_bar = QLabel("Listo")
        self.status_bar.setStyleSheet("color: #7f8c8d; padding: 5px;")
        layout.addWidget(self.status_bar)
        
        # Botón cerrar
        btn_cerrar = QPushButton("CERRAR")
        btn_cerrar.clicked.connect(self.accept)
        btn_cerrar.setObjectName("btn_cerrar")
        layout.addWidget(btn_cerrar)

    def _crear_toolbar(self):
        toolbar = QHBoxLayout()
        
        self.titulo_label = QLabel("📊 GUÍAS PROCESADAS")
        self.titulo_label.setFont(QFont("Arial", 8, QFont.Weight.Bold))
        toolbar.addWidget(self.titulo_label)
        
        toolbar.addStretch()
        
        toolbar.addWidget(QLabel("Filtrar:"))
        self.filtro_combo = QComboBox()
        self.filtro_combo.addItems(["Todos", "✅ Exitosas", "📦 ENT", "❌ Errores", "⚠️ Advertencias"])
        self.filtro_combo.currentTextChanged.connect(self.aplicar_filtro)
        toolbar.addWidget(self.filtro_combo)
        
        self.btn_limpiar = QPushButton("🗑️ Limpiar filtros")
        self.btn_limpiar.clicked.connect(self.limpiar_filtros)
        self.btn_limpiar.setObjectName("btn_limpiar")
        toolbar.addWidget(self.btn_limpiar)
        
        self.btn_exportar_csv = QPushButton("📥 EXPORTAR CSV")
        self.btn_exportar_csv.clicked.connect(self.exportar_csv)
        self.btn_exportar_csv.setObjectName("btn_exportar_csv")
        toolbar.addWidget(self.btn_exportar_csv)
        
        self.btn_exportar_excel = QPushButton("📊 EXPORTAR EXCEL")
        self.btn_exportar_excel.clicked.connect(self.exportar_excel)
        self.btn_exportar_excel.setObjectName("btn_exportar_excel")
        toolbar.addWidget(self.btn_exportar_excel)
        
        return toolbar

    def _crear_tabla(self):
        tabla = QTableWidget()
        tabla.setColumnCount(5)
        tabla.setHorizontalHeaderLabels(["Guía", "Estado", "Resultado", "Navegador", "Fecha/Hora"])
        tabla.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        tabla.setAlternatingRowColors(True)
        tabla.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        tabla.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        tabla.itemDoubleClicked.connect(self.copiar_guia)
        return tabla

    def _setup_styles(self):
        self.setStyleSheet("""
            QDialog {
                background-color:#0f0f0f;
            }
            QPushButton#btn_limpiar {
                background-color: #95a5a6;
                color: white;
                padding: 5px 15px;
                border-radius: 3px;
            }
            QPushButton#btn_limpiar:hover {
                background-color: #7f8c8d;
            }
            QPushButton#btn_exportar_csv {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 5px;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton#btn_exportar_csv:hover {
                background-color: #2ecc71;
            }
            QPushButton#btn_exportar_excel {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 8px 20px;
                border-radius: 5px;
                font-size: 11pt;
                min-width: 150px;
            }
            QPushButton#btn_exportar_excel:hover {
                background-color: #2980b9;
            }
            QPushButton#btn_cerrar {
                background-color: #3498db;
                color: white;
                font-weight: bold;
                padding: 10px;
                border-radius: 5px;
                min-width: 100px;
            }
            QPushButton#btn_cerrar:hover {
                background-color: #2980b9;
            }
        """)

    def actualizar_historial(self, datos):
        self.datos_completos = datos.copy()
        self.datos_filtrados = datos.copy()
        self._aplicar_filtro_actual()
        self._actualizar_vista()

    def _actualizar_vista(self):
        datos_ordenados = sorted(self.datos_filtrados, key=lambda x: x[4], reverse=True)
        self.tabla.setRowCount(len(datos_ordenados))
        self.tabla.setSortingEnabled(False)
        
        for i, (guia, estado, resultado, nav, fecha) in enumerate(datos_ordenados):
            self._agregar_fila(i, guia, estado, resultado, nav, fecha)
        
        self._ajustar_columnas()
        self.tabla.setSortingEnabled(True)

    def _agregar_fila(self, fila, guia, estado, resultado, nav, fecha):
        # Guía
        item_guia = QTableWidgetItem(guia)
        item_guia.setToolTip(f"Haz doble clic para copiar: {guia}")
        self.tabla.setItem(fila, 0, item_guia)
        
        # Estado
        item_estado = QTableWidgetItem(estado)
        config_estado = self._get_estado_config(estado)
        item_estado.setForeground(QColor(config_estado['color']))
        item_estado.setBackground(QColor(config_estado['background']))
        item_estado.setToolTip(config_estado['tooltip'])
        font = QFont()
        font.setBold(True)
        item_estado.setFont(font)
        self.tabla.setItem(fila, 1, item_estado)
        
        # Resultado
        item_resultado = QTableWidgetItem(resultado)
        config_resultado = self._get_resultado_config(resultado)
        item_resultado.setForeground(QColor(config_resultado['color']))
        item_resultado.setBackground(QColor(config_resultado['background']))
        item_resultado.setToolTip(config_resultado['tooltip'])
        self.tabla.setItem(fila, 2, item_resultado)
        
        # Navegador
        item_nav = QTableWidgetItem(nav)
        item_nav.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tabla.setItem(fila, 3, item_nav)
        
        # Fecha
        item_fecha = QTableWidgetItem(fecha)
        item_fecha.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.tabla.setItem(fila, 4, item_fecha)

    def _get_estado_config(self, estado):
        configs = {
            "✅": {'color': "#27ae60", 'background': "#e8f8f5", 'tooltip': "✅ Procesada exitosamente"},
            "📦": {'color': "#f39c12", 'background': "#fff3cd", 'tooltip': "📦 Guía entregada (ENT)"},
            "❌": {'color': "#e74c3c", 'background': "#fdeded", 'tooltip': "❌ Error en procesamiento"},
            "⚠️": {'color': "#f39c12", 'background': "#fff3cd", 'tooltip': "⚠️ Advertencia - Verificar"},
            "⏭️": {'color': "#7f8c8d", 'background': "#ecf0f1", 'tooltip': "⏭️ Omitida - Ya procesada"}
        }
        for key, config in configs.items():
            if key in estado:
                return config
        return {'color': "#7f8c8d", 'background': "#ecf0f1", 'tooltip': "Estado desconocido"}

    def _get_resultado_config(self, resultado):
        if "ENT" in resultado:
            return {'color': "#f39c12", 'background': "#fff3cd", 'tooltip': "📦 Guía con estado ENT"}
        elif "ADVERTENCIA" in resultado or "NO CONFIRMADO" in resultado:
            return {'color': "#f39c12", 'background': "#fff3cd", 'tooltip': "⚠️ Completado con advertencias"}
        elif "ERROR" in resultado:
            return {'color': "#e74c3c", 'background': "#fdeded", 'tooltip': "❌ Error en procesamiento"}
        elif "COMPLETADO" in resultado:
            return {'color': "#27ae60", 'background': "#e8f8f5", 'tooltip': "✅ Procesado correctamente"}
        elif "SIN RESULTADOS" in resultado:
            return {'color': "#e74c3c", 'background': "#fdeded", 'tooltip': "❌ Guía no encontrada"}
        return {'color': "#7f8c8d", 'background': "#ecf0f1", 'tooltip': resultado}

    def _ajustar_columnas(self):
        self.tabla.resizeColumnsToContents()
        self.tabla.setColumnWidth(0, 150)
        self.tabla.setColumnWidth(1, 150)
        self.tabla.setColumnWidth(2, 200)
        self.tabla.setColumnWidth(3, 100)
        self.tabla.setColumnWidth(4, 150)

    def _aplicar_filtro_actual(self):
        if self.filtro_actual == "Todos":
            self.datos_filtrados = self.datos_completos.copy()
            return
        
        mapa_filtro = {
            "✅ Exitosas": "✅",
            "📦 ENT": "📦",
            "❌ Errores": "❌",
            "⚠️ Advertencias": "⚠️"
        }
        estado_filtro = mapa_filtro.get(self.filtro_actual, "")
        self.datos_filtrados = [d for d in self.datos_completos if estado_filtro in d[1]]

    def aplicar_filtro(self, filtro):
        if not self.datos_completos:
            self.status_bar.setText("⚠️ No hay datos para filtrar")
            return
        self.filtro_actual = filtro
        self._aplicar_filtro_actual()
        self._actualizar_vista()

    def limpiar_filtros(self):
        self.filtro_combo.setCurrentText("Todos")
        self.filtro_actual = "Todos"
        self.datos_filtrados = self.datos_completos.copy()
        self._actualizar_vista()
        self.status_bar.setText("✅ Filtros limpiados")

    def copiar_guia(self, item):
        if item.column() == 0:
            guia = item.text()
            QApplication.clipboard().setText(guia)
            item.setSelected(True)
            self.status_bar.setText(f"✅ Guía '{guia}' copiada al portapapeles")
            QTimer.singleShot(3000, lambda: self.status_bar.setText("Listo"))

    def exportar_csv(self):
        try:
            if not self.datos_filtrados:
                QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filtro_text = self.filtro_actual.lower().replace(" ", "_")
            base_nombre = f"historial_alertran_{filtro_text}_{timestamp}"
            
            ruta_completa = self.file_utils.generar_nombre_unico(
                self.carpeta_descargas, base_nombre, "csv"
            )
            
            with open(ruta_completa, 'w', encoding='utf-8-sig') as f:
                f.write("Guía,Estado,Resultado,Navegador,Fecha\n")
                for i in range(self.tabla.rowCount()):
                    fila = [
                        self.tabla.item(i, 0).text(),
                        self.tabla.item(i, 1).text(),
                        self.tabla.item(i, 2).text(),
                        self.tabla.item(i, 3).text(),
                        self.tabla.item(i, 4).text()
                    ]
                    f.write(','.join(fila) + '\n')
            
            QMessageBox.information(self, "✅ Exportación Exitosa", f"Archivo guardado en:\n{ruta_completa}")
            self.status_bar.setText(f"✅ Exportado: {Path(ruta_completa).name}")
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo exportar:\n{str(e)}")

    def exportar_excel(self):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            
            if not self.datos_filtrados:
                QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
                return
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filtro_text = self.filtro_actual.lower().replace(" ", "_")
            base_nombre = f"historial_alertran_{filtro_text}_{timestamp}"
            
            ruta_completa = self.file_utils.generar_nombre_unico(
                self.carpeta_descargas, base_nombre, "xlsx"
            )
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial Alertran"
            
            headers = ["Guía", "Estado", "Resultado", "Navegador", "Fecha"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
            
            for row in range(self.tabla.rowCount()):
                for col in range(self.tabla.columnCount()):
                    item = self.tabla.item(row, col)
                    if item:
                        ws.cell(row=row + 2, column=col + 1, value=item.text())
            
            wb.save(ruta_completa)
            QMessageBox.information(self, "✅ Exportación Exitosa", f"Archivo guardado en:\n{ruta_completa}")
            self.status_bar.setText(f"✅ Exportado: {Path(ruta_completa).name}")
        except Exception as e:
            QMessageBox.critical(self, "❌ Error", f"No se pudo exportar a Excel:\n{str(e)}")